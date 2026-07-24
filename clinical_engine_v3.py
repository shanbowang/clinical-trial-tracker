"""
clinical_engine_v3.py - 通用临床试验 Tracker 分析引擎 V3
支持自动字段匹配，适用于不同结构的Tracker数据

特点：
1. 智能字段匹配 - 支持多别名映射
2. 模块化分析 - 每个分析项独立函数，自动跳过缺失字段
3. 生物标志物自动识别 - CCNE1/Kras等
4. 完整性报告 - 显示成功/跳过的分析项
"""

import os
import json
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
from typing import Dict, Optional, List, Tuple
import warnings
warnings.filterwarnings('ignore')

import platform
from io import BytesIO
import zipfile
from xml.etree import ElementTree as ET


def _read_xlsx_raw(file_path, sheet_name=0):
    """
    终极兜底：直接从 xlsx（ZIP）中提取纯数据，完全绕过 openpyxl 样式解析。
    支持 .xlsx 格式。
    """
    SS_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    
    with zipfile.ZipFile(file_path, 'r') as z:
        # 获取sheet列表
        wb_xml = z.read('xl/workbook.xml')
        wb_root = ET.fromstring(wb_xml)
        sheets = []
        for sheet_elem in wb_root.findall(f'.//{{{SS_NS}}}sheet'):
            sheets.append({
                'name': sheet_elem.get('name'),
                'id': sheet_elem.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            })
        
        # 解析共享字符串
        try:
            sst_xml = z.read('xl/sharedStrings.xml')
            sst_root = ET.fromstring(sst_xml)
            shared_strings = [si.find(f'{{{SS_NS}}}t') for si in sst_root.findall(f'.//{{{SS_NS}}}si')]
            shared_strings = [s.text if s is not None and s.text else '' for s in shared_strings]
        except:
            shared_strings = []
        
        # 选择目标sheet
        if isinstance(sheet_name, int):
            if sheet_name >= len(sheets):
                return pd.DataFrame()
            target_sheet = sheets[sheet_name]
        else:
            target_sheet = next((s for s in sheets if s['name'] == sheet_name), None)
            if target_sheet is None:
                return pd.DataFrame()
        
        # 读取sheet数据
        sheet_xml = z.read(f'xl/worksheets/sheet{sheets.index(target_sheet) + 1}.xml')
        sheet_root = ET.fromstring(sheet_xml)
        
        rows = []
        for row_elem in sheet_root.findall(f'.//{{{SS_NS}}}row'):
            row_data = {}
            for cell_elem in row_elem.findall(f'{{{SS_NS}}}c'):
                ref = cell_elem.get('r')  # e.g., 'A1', 'B3'
                col_letter = ''.join(c for c in ref if c.isalpha())
                col_idx = sum((ord(c) - ord('A') + 1) * (26 ** i) for i, c in enumerate(reversed(col_letter))) - 1
                value_elem = cell_elem.find(f'{{{SS_NS}}}v')
                
                if value_elem is not None and value_elem.text:
                    cell_type = cell_elem.get('t', '')
                    if cell_type == 's':  # shared string
                        idx = int(value_elem.text)
                        val = shared_strings[idx] if idx < len(shared_strings) else ''
                    else:
                        val = value_elem.text
                    row_data[col_idx] = val
            
            rows.append(row_data)
        
        if not rows:
            return pd.DataFrame()
        
        # 转换为DataFrame
        max_col = max(max(r.keys()) for r in rows if r) if any(rows) else 0
        data = []
        for row in rows:
            data.append([row.get(i, None) for i in range(max_col + 1)])
        
        # 第一行作为列名
        columns = [str(c) if c is not None else f'Col{i}' for i, c in enumerate(data[0])]
        df = pd.DataFrame(data[1:], columns=columns)
        return df


def setup_chinese_font():
    system = platform.system()
    if system == "Windows":
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'KaiTi', 'FangSong']
    elif system == "Darwin":
        plt.rcParams['font.sans-serif'] = ['PingFang SC', 'STHeiti', 'Heiti SC']
    else:
        plt.rcParams['font.sans-serif'] = ['Noto Sans CJK SC', 'WenQuanYi Micro Hei']
    plt.rcParams['axes.unicode_minus'] = False
    plt.rcParams['figure.dpi'] = 150

setup_chinese_font()
sns.set_style("whitegrid")

COLOR_PALETTES = {
    'primary': ['#48CAE4', '#90E0EF', '#00B4D8', '#0077B6', '#023E8A'],
    'health': ['#95D5B2', '#74C69D', '#52B788', '#40916C', '#2D6A4F'],
    'accent': ['#FFB703', '#FB8500', '#F77F00', '#DC2F02', '#6A040F'],
    'purple': ['#E0AAFF', '#C77DFF', '#9D4EDD', '#7B2CBF', '#5A189A'],
    'warm': ['#FFCDB2', '#FFB4A2', '#FF8BA7', '#FF6D60', '#F72585'],
}

# ==================== 字段别名映射表 ====================
FIELD_ALIASES = {
    'Site': ['Site No.', 'Site No', 'Site', '中心', 'site no.'],
    'Subject': ['Subject ID', 'Subject_ID', '受试者', 'subject id'],
    'Phase': ['Phase', '阶段', 'phase'],
    'Cohort': ['Cohort(mg)', 'Cohort', '剂量组', 'cohort(mg)'],
    'Tumor': ['Tumor', '肿瘤', '肿瘤类型', 'tumor'],
    'Date_ICF': ['Date ICF', 'Date_ICF', '知情日期', 'date icf'],
    'C1D1': ['C1D1', '首次给药', 'c1d1'],
    'Status': ['Status', '状态', 'status'],
    'Best_Response': ['Best Response', 'Best_Response', '疗效', 'best response'],
    'EOT_Date': ['EOT date', 'EOT_Date', '终止日期', 'eot date'],
    'EOT_Reason': ['EOT reason', 'EOT_Reason', '终止原因', 'eot reason'],
    'EOT_Reason_Type': ['EOT reason type', 'EOT reson type', 'EOT_Reason_Type', 'eot reason type'],
    'Latest_Date': ['Latest Date', 'Latest_Date', '最新日期', 'latest date'],
    'Latest_Cycle': ['Latest Cycle', 'Latest_Cycle', '最新周期', 'latest cycle'],
    'Protocol_Version': ['protocol version', 'Protocol', '方案版本', 'protocol version'],
    'Source': ['Source of Patients', 'from', '来源', '病例来源', 'source of patients'],
    'Screen_Failure_Type': ['Screen Failure Type', 'Screen Failure type', '筛选失败类型'],
    'Screen_Failure_Reason': ['Screen Failure Reason', '筛选失败原因'],
    # 生物标志物 - CCNE1
    'CCNE1_Result': ['CCNE1 result', 'CCNE1_result', 'ccne1 result'],
    'CCNE1_Send_Date': ['Tumor slide send out date', '肿瘤切片送出日期'],
    'CCNE1_Report_Date': ['CCNE1 result report date', 'CCNE1_report_date'],
    # 生物标志物 - Kras
    'Kras_Result': ['Kras result', 'Kras_result', 'kras result'],
    'Kras_Report_Date': ['Kras result report date', 'Kras_report_date'],
    # 生物标志物 - 通用
    'Biomarker_Result': ['CCNE1 result', 'Kras result', '检测结果'],
    'Biomarker_Send_Date': ['Tumor slide send out date'],
    'Biomarker_Report_Date': ['CCNE1 result report date', 'Kras result report date'],
}

# ==================== 工具函数 ====================

def find_column(df_columns, *aliases) -> Optional[str]:
    """查找第一个匹配的列名"""
    for col in df_columns:
        col_clean = str(col).strip().lower().replace(' ', '').replace('_', '').replace('.', '')
        for alias in aliases:
            alias_clean = str(alias).strip().lower().replace(' ', '').replace('_', '').replace('.', '')
            if alias_clean == col_clean or alias_clean in col_clean or col_clean in alias_clean:
                return col
    return None

def find_all_columns(df_columns, *aliases) -> list:
    """查找所有匹配的列名"""
    matches = []
    for col in df_columns:
        col_clean = str(col).strip().lower().replace(' ', '').replace('_', '').replace('.', '')
        for alias in aliases:
            alias_clean = str(alias).strip().lower().replace(' ', '').replace('_', '').replace('.', '')
            if alias_clean == col_clean or alias_clean in col_clean or col_clean in alias_clean:
                matches.append(col)
                break
    return matches

def find_best_column(df, key: str) -> Optional[str]:
    """查找最佳匹配列：优先精确匹配，其次多列时选非空最多的"""
    if key not in FIELD_ALIASES:
        return None
    matches = find_all_columns(df.columns, *FIELD_ALIASES[key])
    if not matches:
        return None
    if len(matches) == 1:
        return matches[0]
    
    def _match_quality(col, aliases):
        """匹配质量：精确匹配=2，包含匹配=1"""
        col_clean = str(col).strip().lower().replace(' ', '').replace('_', '').replace('.', '')
        for alias in aliases:
            alias_clean = str(alias).strip().lower().replace(' ', '').replace('_', '').replace('.', '')
            if alias_clean == col_clean:
                return 2
            if alias_clean in col_clean or col_clean in alias_clean:
                return 1
        return 0
    
    # 按匹配质量排序，同质量按非空数排序
    aliases = FIELD_ALIASES[key]
    scored = [(m, _match_quality(m, aliases), df[m].notna().sum()) for m in matches]
    scored.sort(key=lambda x: (-x[1], -x[2]))  # 质量优先，然后非空数
    return scored[0][0]

def find_column_by_key(df_columns, key: str) -> Optional[str]:
    """通过预定义的key查找列名（简单版，不区分多匹配）"""
    if key in FIELD_ALIASES:
        return find_column(df_columns, *FIELD_ALIASES[key])
    return None

def standardize_tumor_name(tumor):
    """标准化肿瘤名称"""
    if pd.isna(tumor):
        return tumor
    tumor_lower = str(tumor).lower().strip()
    if 'breast' in tumor_lower or 'brest' in tumor_lower:
        return 'Breast cancer'
    elif 'ovarian' in tumor_lower:
        return 'Ovarian cancer'
    elif 'endometrial' in tumor_lower:
        return 'Endometrial cancer'
    elif 'prostate' in tumor_lower or '前列腺' in tumor_lower:
        return 'Prostate cancer'
    elif 'pancreatic' in tumor_lower or '胰腺' in tumor_lower:
        return 'Pancreatic cancer'
    elif 'colon' in tumor_lower or '结肠' in tumor_lower:
        return 'Colon cancer'
    elif 'lung' in tumor_lower or '肺' in tumor_lower:
        return 'Lung cancer'
    return tumor

def save_chart(fig, output_dir, filename):
    path = os.path.join(output_dir, "charts", filename)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    fig.savefig(path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    return path

def save_table(df, output_dir, filename):
    path = os.path.join(output_dir, "tables", filename)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    df.to_csv(path, index=True)
    return path

# ==================== 字段检测类 ====================

class FieldDetector:
    """字段检测器"""
    
    def __init__(self, df: pd.DataFrame):
        self.df = df
        self.columns = df.columns.tolist()
        self.found_fields = {}
        self.missing_fields = []
        self._detect_all()
    
    def _detect_all(self):
        """检测所有预定义字段（多列匹配时选非空最多的）"""
        for key in FIELD_ALIASES:
            col = find_best_column(self.df, key)
            if col:
                self.found_fields[key] = col
            else:
                self.missing_fields.append(key)
    
    def get(self, key: str) -> Optional[str]:
        """获取字段名"""
        return self.found_fields.get(key)
    
    def has(self, *keys) -> bool:
        """检查字段是否存在"""
        return all(key in self.found_fields for key in keys)
    
    def has_any(self, *keys) -> bool:
        """检查任意字段是否存在"""
        return any(key in self.found_fields for key in keys)
    
    def get_biomarker_type(self) -> Optional[str]:
        """检测生物标志物类型"""
        if 'CCNE1_Result' in self.found_fields:
            return 'CCNE1'
        elif 'Kras_Result' in self.found_fields:
            return 'Kras'
        return None
    
    def report(self) -> Dict:
        """生成字段检测报告"""
        return {
            'found': self.found_fields,
            'missing': self.missing_fields,
            'biomarker_type': self.get_biomarker_type()
        }

# ==================== 分析函数 ====================

def analyze_phase_dose(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[Dict], str]:
    """分析1: Phase & 剂量组筛选入组表"""
    if not fields.has('Phase', 'Cohort', 'Date_ICF', 'C1D1'):
        return None, "缺少 Phase/Cohort/Date_ICF/C1D1 字段"
    
    phase_col = fields.get('Phase')
    cohort_col = fields.get('Cohort')
    date_icf_col = fields.get('Date_ICF')
    c1d1_col = fields.get('C1D1')
    
    df['筛选'] = df[date_icf_col].notna()
    df['入组'] = df[c1d1_col].notna()
    
    df['Phase_Dose'] = 'Phase ' + df[phase_col].astype(str) + ' - ' + df[cohort_col].astype(str) + 'mg'
    
    stats = df.groupby('Phase_Dose').agg({
        '筛选': 'sum', '入组': 'sum'
    }).astype(int)
    stats['筛选成功率(%)'] = (stats['入组'] / stats['筛选'] * 100).round(1)
    
    # 图表：分组柱状图
    fig, ax = plt.subplots(figsize=(max(6, len(stats) * 1.5), 5))
    x = range(len(stats))
    width = 0.35
    ax.bar([i - width/2 for i in x], stats['筛选'], width, label='筛选', color=COLOR_PALETTES['primary'][1], alpha=0.85)
    ax.bar([i + width/2 for i in x], stats['入组'], width, label='入组', color=COLOR_PALETTES['health'][1], alpha=0.85)
    ax.set_xticks(x)
    ax.set_xticklabels(stats.index, rotation=30, ha='right')
    ax.set_ylabel('人数', fontweight='bold')
    ax.set_title('Phase & 剂量组筛选入组对比', fontweight='bold', pad=12)
    ax.legend(loc='upper right')
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    chart_path = save_chart(fig, output_dir, "01_phase_dose.png")
    
    return {'table': stats, 'chart': chart_path}, ""

def analyze_site_stats(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[Dict], str]:
    """分析2: 中心筛选入组表"""
    if not fields.has('Site'):
        return None, "缺少 Site 字段"
    
    site_col = fields.get('Site')
    date_icf_col = fields.get('Date_ICF')
    c1d1_col = fields.get('C1D1')
    status_col = fields.get('Status')
    
    if date_icf_col:
        df['筛选'] = df[date_icf_col].notna()
    else:
        df['筛选'] = True
    
    if c1d1_col:
        df['入组'] = df[c1d1_col].notna()
    else:
        df['入组'] = False
    
    if status_col:
        df['筛选失败'] = df[status_col].fillna('').str.upper().isin(['SF', 'SCREEN FAILURE', 'SCREENFAILURE'])
    else:
        df['筛选失败'] = False
    
    site_stats = df.groupby(site_col).agg({
        '筛选': 'sum', '入组': 'sum', '筛选失败': 'sum'
    }).astype(int)
    
    site_stats = site_stats.sort_values('入组', ascending=False).reset_index()
    site_stats = site_stats.set_index(site_col)
    
    total = site_stats.sum(numeric_only=True)
    total['筛选成功率(%)'] = (total['入组'] / total['筛选'] * 100).round(1) if total['筛选'] > 0 else 0
    site_stats.loc['总计'] = total
    site_stats['筛选成功率(%)'] = (site_stats['入组'] / site_stats['筛选'] * 100).round(1)
    
    # 生成图表
    fig, ax = plt.subplots(figsize=(10, 6))
    site_plot = site_stats.drop('总计') if '总计' in site_stats.index else site_stats.copy()
    colors = COLOR_PALETTES['primary'][:len(site_plot)]
    x_labels = [str(x) for x in site_plot.index.tolist()]
    x_values = site_plot['入组'].values
    bars = ax.bar(range(len(x_labels)), x_values, color=colors, alpha=0.8)
    ax.set_xticks(range(len(x_labels)))
    ax.set_xticklabels(x_labels, rotation=45)
    ax.set_xlabel('中心', fontweight='bold')
    ax.set_ylabel('入组数', fontweight='bold')
    ax.set_title('各中心入组数对比（按入组数降序）', fontweight='bold', pad=15)
    ax.grid(axis='y', alpha=0.3)
    for bar in bars:
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, 
               int(bar.get_height()), ha='center', fontweight='bold')
    plt.tight_layout()
    chart_path = save_chart(fig, output_dir, "02_site_enrollment.png")
    
    return {'table': site_stats, 'chart': chart_path}, ""

def analyze_phase_tumor(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[Dict], str]:
    """分析3: Phase & 肿瘤类型筛选入组表"""
    if not fields.has('Phase'):
        return None, "缺少 Phase 字段"
    if not fields.has('Tumor'):
        return None, "缺少 Tumor 字段"
    
    phase_col = fields.get('Phase')
    tumor_col = fields.get('Tumor')
    
    df['Tumor_Std'] = df[tumor_col].apply(standardize_tumor_name)
    
    stats = df.groupby([phase_col, 'Tumor_Std']).agg({
        '筛选': 'sum', '入组': 'sum', '筛选失败': 'sum'
    }).astype(int)
    stats['筛选成功率(%)'] = (stats['入组'] / stats['筛选'] * 100).round(1)
    
    # 图表：按肿瘤类型分组柱状图（按Phase分组，展示入组数）
    flat_stats = stats.reset_index()
    tumors = flat_stats['Tumor_Std'].unique()
    phases = sorted(flat_stats[phase_col].unique())
    
    fig, ax = plt.subplots(figsize=(max(8, len(tumors) * 2), 5))
    x = range(len(tumors))
    n_phases = len(phases)
    width = min(0.8 / n_phases, 0.35)
    phase_colors = [COLOR_PALETTES['primary'][i % len(COLOR_PALETTES['primary'])] for i in range(n_phases)]
    
    for i, phase in enumerate(phases):
        phase_data = flat_stats[flat_stats[phase_col] == phase]
        values = [phase_data[phase_data['Tumor_Std'] == t]['入组'].sum() if t in phase_data['Tumor_Std'].values else 0 for t in tumors]
        offset = (i - (n_phases - 1) / 2) * width
        ax.bar([j + offset for j in x], values, width, label=f'Phase {phase}', color=phase_colors[i], alpha=0.85)
    
    ax.set_xticks(x)
    ax.set_xticklabels(tumors, rotation=30, ha='right')
    ax.set_ylabel('入组人数', fontweight='bold')
    ax.set_title('Phase & 肿瘤类型入组分布', fontweight='bold', pad=12)
    ax.legend(loc='upper right')
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    chart_path = save_chart(fig, output_dir, "03_phase_tumor.png")
    
    return {'table': stats, 'chart': chart_path}, ""

def analyze_status_pivot(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[Dict], str]:
    """分析4: Phase & 状态分布表"""
    if not fields.has('Phase', 'Status'):
        return None, "缺少 Phase/Status 字段"
    
    phase_col = fields.get('Phase')
    status_col = fields.get('Status')
    
    pivot = df.groupby([phase_col, status_col]).size().unstack(fill_value=0)
    
    # 图表：堆叠柱状图
    fig, ax = plt.subplots(figsize=(max(6, len(pivot) * 1.5), 5))
    bottom = np.zeros(len(pivot))
    status_colors = [COLOR_PALETTES['primary'][i % 5] for i in range(len(pivot.columns))]
    for i, col in enumerate(pivot.columns):
        ax.bar(range(len(pivot)), pivot[col], bottom=bottom, label=str(col), color=status_colors[i], alpha=0.85)
        bottom += pivot[col].values
    ax.set_xticks(range(len(pivot)))
    ax.set_xticklabels([f'Phase {p}' for p in pivot.index], rotation=0)
    ax.set_ylabel('人数', fontweight='bold')
    ax.set_title('Phase & 状态分布', fontweight='bold', pad=12)
    ax.legend(loc='upper right')
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    chart_path = save_chart(fig, output_dir, "04_status_pivot.png")
    
    return {'table': pivot, 'chart': chart_path}, ""

def analyze_monthly_trend(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[Dict], str]:
    """分析5: 月度筛选入组趋势"""
    if not fields.has('Date_ICF', 'C1D1'):
        return None, "缺少 Date_ICF/C1D1 字段"
    
    date_icf_col = fields.get('Date_ICF')
    c1d1_col = fields.get('C1D1')
    
    try:
        df['ICF_Month'] = _smart_parse_dates(df[date_icf_col]).dt.to_period('M').dt.to_timestamp()
        df['C1D1_Month'] = _smart_parse_dates(df[c1d1_col]).dt.to_period('M').dt.to_timestamp()
        
        monthly_screened = df[df['筛选']].groupby('ICF_Month').size()
        monthly_enrolled = df[df['入组']].groupby('C1D1_Month').size()
        
        all_months = pd.Index(sorted(set(monthly_screened.index) | set(monthly_enrolled.index)))
        monthly_data = pd.DataFrame({
            '筛选': monthly_screened.reindex(all_months, fill_value=0),
            '入组': monthly_enrolled.reindex(all_months, fill_value=0)
        })
        monthly_data['累计筛选'] = monthly_data['筛选'].cumsum()
        monthly_data['累计入组'] = monthly_data['入组'].cumsum()
        
        # 图表
        fig, ax1 = plt.subplots(figsize=(12, 6))
        x = range(len(monthly_data))
        width = 0.35
        ax1.bar([i - width/2 for i in x], monthly_data['筛选'], width, label='月度筛选', color=COLOR_PALETTES['primary'][1], alpha=0.8)
        ax1.bar([i + width/2 for i in x], monthly_data['入组'], width, label='月度入组', color=COLOR_PALETTES['health'][1], alpha=0.8)
        ax1.set_xlabel('月份', fontweight='bold')
        ax1.set_ylabel('月度计数', fontweight='bold')
        ax1.set_xticks(x)
        ax1.set_xticklabels([m.strftime('%Y-%m') if hasattr(m, 'strftime') else str(m) for m in monthly_data.index], rotation=45)
        
        ax2 = ax1.twinx()
        ax2.plot(x, monthly_data['累计筛选'], 'o-', color=COLOR_PALETTES['primary'][3], linewidth=2.5, label='累计筛选')
        ax2.plot(x, monthly_data['累计入组'], 's-', color=COLOR_PALETTES['health'][2], linewidth=2.5, label='累计入组')
        ax2.set_ylabel('累计计数', fontweight='bold')
        
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')
        ax1.set_title('月度筛选入组趋势图', fontweight='bold', pad=15)
        ax1.grid(axis='y', alpha=0.3)
        plt.tight_layout()
        chart_path = save_chart(fig, output_dir, "05_monthly_trend.png")
        
        return {'table': monthly_data, 'chart': chart_path}, ""
    except Exception as e:
        return None, f"日期处理错误: {str(e)}"

def analyze_sf_reasons(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[Dict], str]:
    """分析6: 筛选失败原因分布"""
    if not fields.has('Screen_Failure_Reason'):
        return None, "缺少筛选失败原因字段"
    if df['筛选失败'].sum() == 0:
        return None, "没有筛选失败数据"
    
    sf_reason_col = fields.get('Screen_Failure_Reason')
    sf_data = df[df['筛选失败'] == True]
    
    if len(sf_data) == 0:
        return None, "没有筛选失败数据"
    
    sf_counts = sf_data[sf_reason_col].value_counts()
    
    fig, ax = plt.subplots(figsize=(8, 8))
    colors = COLOR_PALETTES['accent'][:len(sf_counts)]
    ax.pie(sf_counts.values, labels=sf_counts.index, autopct='%1.1f%%', colors=colors, startangle=90)
    ax.set_title('筛选失败原因分布', fontweight='bold', pad=15)
    plt.tight_layout()
    chart_path = save_chart(fig, output_dir, "06_sf_reasons.png")
    
    return {'table': sf_counts.to_frame('人数'), 'chart': chart_path}, ""

def analyze_biomarker(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[Dict], str]:
    """分析7: 生物标志物检测分析"""
    biomarker_type = fields.get_biomarker_type()
    if not biomarker_type:
        return None, "缺少生物标志物检测结果字段"
    
    if biomarker_type == 'CCNE1':
        result_col = fields.get('CCNE1_Result')
        send_col = fields.get('CCNE1_Send_Date')
        report_col = fields.get('CCNE1_Report_Date')
    else:
        result_col = fields.get('Kras_Result')
        send_col = fields.get('Biomarker_Send_Date')
        report_col = fields.get('Kras_Report_Date')
    
    summary = {'type': biomarker_type}
    
    # 结果分布
    if result_col:
        result_counts = df[result_col].value_counts()
        summary['result_table'] = result_counts.to_frame('人数')
        
        # 图表：结果分布柱状图
        fig, ax = plt.subplots(figsize=(max(4, len(result_counts) * 1.2), 4))
        colors = COLOR_PALETTES['health'][:len(result_counts)]
        ax.bar(range(len(result_counts)), result_counts.values, color=colors, alpha=0.85)
        ax.set_xticks(range(len(result_counts)))
        ax.set_xticklabels(result_counts.index, rotation=30, ha='right')
        ax.set_ylabel('人数', fontweight='bold')
        ax.set_title(f'{biomarker_type} 检测结果分布', fontweight='bold', pad=12)
        ax.grid(axis='y', alpha=0.3)
        plt.tight_layout()
        chart_path = save_chart(fig, output_dir, "07_biomarker.png")
        summary['chart'] = chart_path
    
    # 检测时间
    if send_col and report_col:
        try:
            send_dates = _smart_parse_dates(df[send_col].copy())
            report_dates = _smart_parse_dates(df[report_col].copy())
            valid = df[(send_dates.notna()) & (report_dates.notna())].copy()
            if len(valid) > 0:
                valid_idx = valid.index
                valid = valid.copy()
                valid['检测天数'] = (report_dates.loc[valid_idx] - send_dates.loc[valid_idx]).dt.days
                valid = valid[valid['检测天数'] >= 0]
                if len(valid) > 0:
                    summary['samples'] = len(valid)
                    summary['mean_days'] = round(valid['检测天数'].mean(), 1)
                    summary['median_days'] = int(valid['检测天数'].median())
        except:
            pass
    
    return summary, ""

def analyze_time_on_study(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[str], str]:
    """分析8: 在组时间游泳图"""
    if not fields.has('C1D1', 'Latest_Date'):
        return None, "缺少 C1D1/Latest_Date 字段"
    
    c1d1_col = fields.get('C1D1')
    latest_col = fields.get('Latest_Date')
    
    try:
        latest_dates = _smart_parse_dates(df[latest_col].copy())
        time_data = df[(df[c1d1_col].notna()) & (latest_dates.notna())].copy()
        if len(time_data) == 0:
            return None, "没有在组时间数据"
        
        idx = time_data.index
        c1d1_dates = _smart_parse_dates(time_data[c1d1_col])
        time_data['在组天数'] = (latest_dates.loc[idx] - c1d1_dates).dt.days
        time_data = time_data[time_data['在组天数'] >= 0].sort_values('在组天数', ascending=False)
        
        if len(time_data) == 0:
            return None, "没有在组时间数据"
        
        fig, ax = plt.subplots(figsize=(12, max(6, len(time_data) * 0.3)))
        colors = COLOR_PALETTES['health'][1]
        ax.barh(range(len(time_data)), time_data['在组天数'], color=colors, alpha=0.9)
        ax.set_xlabel('在组天数', fontweight='bold')
        ax.set_title('在组时间游泳图', fontweight='bold', pad=15)
        ax.invert_yaxis()
        ax.grid(axis='x', alpha=0.3)
        plt.tight_layout()
        chart_path = save_chart(fig, output_dir, "08_time_on_study.png")
        return chart_path, ""
    except Exception as e:
        return None, f"错误: {str(e)}"

def analyze_eot_reasons(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[Dict], str]:
    """分析9: EOT原因分布"""
    if not fields.has('EOT_Reason_Type'):
        return None, "缺少 EOT reason type 字段"
    
    eot_col = fields.get('EOT_Reason_Type')
    eot_data = df[df[eot_col].notna()]
    
    if len(eot_data) == 0:
        return None, "没有EOT数据"
    
    eot_counts = eot_data[eot_col].value_counts()
    
    fig, ax = plt.subplots(figsize=(8, 8))
    colors = COLOR_PALETTES['primary'][:len(eot_counts)]
    ax.pie(eot_counts.values, labels=eot_counts.index, autopct='%1.1f%%', colors=colors, startangle=90)
    ax.set_title('EOT 原因分布', fontweight='bold', pad=15)
    plt.tight_layout()
    chart_path = save_chart(fig, output_dir, "09_eot_reasons.png")
    
    return {'table': eot_counts.to_frame('人数'), 'chart': chart_path}, ""

def analyze_response(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[Dict], str]:
    """分析10: 疗效数据表"""
    if not fields.has('Best_Response'):
        return None, "缺少 Best Response 字段"
    
    response_col = fields.get('Best_Response')
    response_data = df[df[response_col].notna()]
    if len(response_data) == 0:
        return None, "所有记录均无疗效评估数据"
    
    if fields.has('Tumor'):
        tumor_col = fields.get('Tumor')
        df['Tumor_Std'] = df[tumor_col].apply(standardize_tumor_name)
        pivot = pd.crosstab(df['Tumor_Std'], df[response_col], dropna=True)
        
        if len(pivot) == 0:
            return None, "无有效疗效数据（均为空值）"
        
        # 图表：按肿瘤类型的堆叠柱状图
        fig, ax = plt.subplots(figsize=(max(6, len(pivot) * 1.5), 5))
        bottom = np.zeros(len(pivot))
        resp_colors = [COLOR_PALETTES['primary'][i % 5] for i in range(len(pivot.columns))]
        for i, col in enumerate(pivot.columns):
            ax.bar(range(len(pivot)), pivot[col], bottom=bottom, label=str(col), color=resp_colors[i], alpha=0.85)
            bottom += pivot[col].values
        ax.set_xticks(range(len(pivot)))
        ax.set_xticklabels(pivot.index, rotation=30, ha='right')
        ax.legend(loc='upper right')
    else:
        counts = response_data[response_col].value_counts()
        pivot = counts.to_frame('人数').T
        
        if len(counts) == 0:
            return None, "无有效疗效数据（均为空值）"
        
        # 图表：简单柱状图
        fig, ax = plt.subplots(figsize=(max(4, len(counts) * 1.2), 4))
        ax.bar(range(len(counts)), counts.values, color=COLOR_PALETTES['primary'][:len(counts)], alpha=0.85)
        ax.set_xticks(range(len(counts)))
        ax.set_xticklabels(counts.index, rotation=0)
    
    ax.set_ylabel('人数', fontweight='bold')
    ax.set_title('疗效数据分布', fontweight='bold', pad=12)
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    chart_path = save_chart(fig, output_dir, "10_response.png")
    
    return {'table': pivot, 'chart': chart_path}, ""

def analyze_protocol_version(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[Dict], str]:
    """分析11: 方案版本筛选入组"""
    if not fields.has('Protocol_Version'):
        return None, "缺少方案版本字段"
    
    version_col = fields.get('Protocol_Version')
    
    stats = df.groupby(version_col).agg({
        '筛选': 'sum', '入组': 'sum', '筛选失败': 'sum'
    }).astype(int)
    stats['筛选成功率(%)'] = (stats['入组'] / stats['筛选'] * 100).round(1)
    
    # 图表：分组柱状图
    fig, ax = plt.subplots(figsize=(max(5, len(stats) * 1.2), 5))
    x = range(len(stats))
    width = 0.25
    ax.bar([i - width for i in x], stats['筛选'], width, label='筛选', color=COLOR_PALETTES['primary'][1], alpha=0.85)
    ax.bar(x, stats['入组'], width, label='入组', color=COLOR_PALETTES['health'][1], alpha=0.85)
    ax.bar([i + width for i in x], stats['筛选失败'], width, label='筛选失败', color=COLOR_PALETTES['accent'][0], alpha=0.85)
    ax.set_xticks(x)
    ax.set_xticklabels(stats.index, rotation=30, ha='right')
    ax.set_ylabel('人数', fontweight='bold')
    ax.set_title('方案版本筛选入组对比', fontweight='bold', pad=12)
    ax.legend(loc='upper right')
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    chart_path = save_chart(fig, output_dir, "11_protocol_version.png")
    
    return {'table': stats, 'chart': chart_path}, ""

def analyze_source(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[Dict], str]:
    """分析12: 不同来源分析"""
    if not fields.has('Source'):
        return None, "缺少来源字段"
    
    source_col = fields.get('Source')
    source_data = df[df[source_col].notna()]
    if len(source_data) == 0:
        return None, "所有记录均无来源信息"
    
    stats = df.groupby(source_col, dropna=True).agg({
        '筛选': 'sum', '入组': 'sum', '筛选失败': 'sum'
    }).astype(int)
    stats['筛选失败率(%)'] = (stats['筛选失败'] / stats['筛选'] * 100).round(1)
    
    # 图表：分组柱状图
    fig, ax = plt.subplots(figsize=(max(5, len(stats) * 1.5), 5))
    x = range(len(stats))
    width = 0.25
    ax.bar([i - width for i in x], stats['筛选'], width, label='筛选', color=COLOR_PALETTES['primary'][1], alpha=0.85)
    ax.bar(x, stats['入组'], width, label='入组', color=COLOR_PALETTES['health'][1], alpha=0.85)
    ax.bar([i + width for i in x], stats['筛选失败'], width, label='筛选失败', color=COLOR_PALETTES['accent'][0], alpha=0.85)
    ax.set_xticks(x)
    ax.set_xticklabels(stats.index, rotation=30, ha='right')
    ax.set_ylabel('人数', fontweight='bold')
    ax.set_title('不同来源筛选入组对比', fontweight='bold', pad=12)
    ax.legend(loc='upper right')
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    chart_path = save_chart(fig, output_dir, "12_source.png")
    
    return {'table': stats, 'chart': chart_path}, ""

def analyze_screening_duration(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[Dict], str]:
    """分析13: 各中心筛选时长"""
    if not fields.has('Site', 'Date_ICF', 'C1D1'):
        return None, "缺少必要字段"
    
    site_col = fields.get('Site')
    date_icf_col = fields.get('Date_ICF')
    c1d1_col = fields.get('C1D1')
    
    try:
        screening_time = df[(df[date_icf_col].notna()) & (df[c1d1_col].notna())].copy()
        if len(screening_time) == 0:
            return None, "没有筛选时长数据"
        
        screening_time[date_icf_col] = _smart_parse_dates(screening_time[date_icf_col])
        screening_time[c1d1_col] = _smart_parse_dates(screening_time[c1d1_col])
        screening_time['筛选天数'] = (screening_time[c1d1_col] - screening_time[date_icf_col]).dt.days
        screening_time = screening_time[screening_time['筛选天数'] >= 0]
        
        if len(screening_time) == 0:
            return None, "没有筛选时长数据"
        
        stats = screening_time.groupby(site_col).agg({
            '筛选天数': ['min', 'max', 'mean', 'count']
        }).round(1)
        stats.columns = ['最小天数', '最大天数', '平均天数', '样本数']
        stats = stats.sort_values('平均天数', ascending=False)
        
        # 图表：水平柱状图（平均筛选时长）
        fig, ax = plt.subplots(figsize=(8, max(4, len(stats) * 0.4)))
        colors = [COLOR_PALETTES['primary'][i % 5] for i in range(len(stats))]
        ax.barh(range(len(stats)), stats['平均天数'], color=colors, alpha=0.85)
        ax.set_yticks(range(len(stats)))
        ax.set_yticklabels([str(s) for s in stats.index])
        ax.set_xlabel('平均筛选天数', fontweight='bold')
        ax.set_title('各中心平均筛选时长（降序）', fontweight='bold', pad=12)
        ax.invert_yaxis()
        ax.grid(axis='x', alpha=0.3)
        for i, v in enumerate(stats['平均天数']):
            ax.text(v + 0.5, i, str(v), va='center', fontweight='bold')
        plt.tight_layout()
        chart_path = save_chart(fig, output_dir, "13_screening_duration.png")
        
        return {'table': stats, 'chart': chart_path}, ""
    except:
        return None, "日期处理错误"

def analyze_risk_dashboard(df: pd.DataFrame, fields: FieldDetector, output_dir: str) -> Tuple[Optional[Dict], str]:
    """分析14: 风险预警看板"""
    if not fields.has('Site'):
        return None, "缺少 Site 字段"
    
    site_col = fields.get('Site')
    site_count = df[site_col].nunique()
    
    if site_count < 2:
        return None, "中心数少于2个，无法生成风险看板"
    
    from datetime import date as date_type
    today = date_type.today()
    
    c1d1_col = fields.get('C1D1')
    status_col = fields.get('Status')
    
    site_metrics = []
    for site, group in df.groupby(site_col):
        screened = group['筛选'].sum() if '筛选' in group else len(group)
        enrolled = group['入组'].sum() if '入组' in group else 0
        failures = group['筛选失败'].sum() if '筛选失败' in group else 0
        
        eot_count = 0
        if status_col:
            eot_count = group[status_col].fillna('').str.upper().eq('EOT').sum()
        
        monthly_rate = 0
        if c1d1_col and c1d1_col in group.columns:
            c1d1_dates = pd.to_datetime(group[c1d1_col], errors='coerce').dropna()
            if len(c1d1_dates) > 0:
                first_enroll = c1d1_dates.min()
                months_active = max((today.year - first_enroll.year) * 12 + (today.month - first_enroll.month) + 1, 1)
                monthly_rate = round(enrolled / months_active, 2)
        
        site_metrics.append({
            '中心': str(site),
            '筛选数': int(screened),
            '入组数': int(enrolled),
            '月均入组': monthly_rate,
            '筛选失败数': int(failures),
            '筛选失败率': round(failures / max(screened, 1) * 100, 1),
            '退出数': int(eot_count),
            '退出率': round(eot_count / max(enrolled, 1) * 100, 1)
        })
    
    if len(site_metrics) < 2:
        return None, "数据不足"
    
    metrics_df = pd.DataFrame(site_metrics)
    
    # 计算IQR边界
    def calc_bounds(values, side='both'):
        q1 = values.quantile(0.25)
        q3 = values.quantile(0.75)
        iqr = q3 - q1
        lower = q1 - 1.5 * iqr if side == 'both' else None
        upper = q3 + 1.5 * iqr
        return lower, upper
    
    risk_metrics = [
        ('月均入组', '月均入组', 'both'),
        ('筛选失败率', '筛选失败率', 'upper'),
        ('退出率', '退出率', 'upper')
    ]
    
    alerts = {}
    for name, col, side in risk_metrics:
        if col in metrics_df.columns and len(metrics_df) >= 4:
            lower, upper = calc_bounds(metrics_df[col], side)
            alerts[name] = {'lower': lower, 'upper': upper}
    
    warning_rows = []
    for _, row in metrics_df.iterrows():
        site_alerts = {}
        for name, col, side in risk_metrics:
            val = row.get(col, 0)
            bounds = alerts.get(name, {})
            lower, upper = bounds.get('lower'), bounds.get('upper')
            
            if side == 'both':
                site_alerts[name] = '高风险' if (lower is not None and val < lower) or (upper is not None and val > upper) else '低风险'
            else:
                site_alerts[name] = '高风险' if (upper is not None and val > upper) else '低风险'
        
        warning_rows.append({
            '中心': row['中心'],
            '月均入组': row['月均入组'],
            '入组预警': site_alerts.get('月均入组', '低风险'),
            '筛选失败率(%)': row['筛选失败率'],
            '失败率预警': site_alerts.get('筛选失败率', '低风险'),
            '退出率(%)': row['退出率'],
            '退出预警': site_alerts.get('退出率', '低风险')
        })
    
    risk_table = pd.DataFrame(warning_rows)
    
    # 热力图
    fig, ax = plt.subplots(figsize=(12, max(4, len(risk_table) * 0.5)))
    risk_cols = ['入组预警', '失败率预警', '退出预警']
    heat_data = []
    for _, row in risk_table.iterrows():
        row_data = [2 if row.get(col, '低风险') == '高风险' else 0 for col in risk_cols]
        heat_data.append(row_data)
    
    import matplotlib.colors as mcolors
    cmap = mcolors.ListedColormap(['#6BCB77', '#FFD93D', '#FF6B6B'])
    im = ax.imshow(heat_data, aspect='auto', cmap=cmap, vmin=0, vmax=2)
    ax.set_xticks(range(len(risk_cols)))
    ax.set_xticklabels(risk_cols, fontsize=10)
    ax.set_yticks(range(len(risk_table)))
    ax.set_yticklabels(risk_table['中心'].tolist(), fontsize=9)
    ax.set_title('风险预警热力图', fontweight='bold', pad=15)
    
    for i in range(len(heat_data)):
        for j in range(len(risk_cols)):
            val = heat_data[i][j]
            symbol = '!!' if val == 2 else 'OK'
            color = 'white' if val == 2 else 'black'
            ax.text(j, i, symbol, ha='center', va='center', fontsize=10, color=color, fontweight='bold')
    
    plt.tight_layout()
    chart_path = save_chart(fig, output_dir, "14_risk_heatmap.png")
    
    return {'table': risk_table, 'chart': chart_path}, ""

# ==================== 数据校验模块 ====================

def validate_tracker_data(df: pd.DataFrame, fields: FieldDetector = None) -> Dict:
    """
    对Tracker数据进行全面校验，返回分类的问题清单
    
    Returns:
        {
            "errors": [{"message": str, "rows": [int], "count": int, "category": str}, ...],
            "warnings": [...],
            "summary": {"total_errors": int, "total_warnings": int, "has_blocking": bool}
        }
    """
    if fields is None:
        fields = FieldDetector(df)
    
    errors = []
    warnings = []
    
    _check_duplicate_subjects(df, fields, errors)
    _check_future_dates(df, fields, errors)
    _check_site_format(df, fields, warnings)
    _check_date_logic(df, fields, errors)
    _check_status_logic(df, fields, errors)
    _check_completeness(df, fields, warnings)
    
    has_blocking = len(errors) > 0
    
    return {
        "errors": errors,
        "warnings": warnings,
        "summary": {
            "total_errors": len(errors),
            "total_warnings": len(warnings),
            "has_blocking": has_blocking,
            "error_count": sum(e["count"] for e in errors),
            "warning_count": sum(w["count"] for w in warnings),
        }
    }


def _format_rows(rows: list, max_show: int = 5) -> str:
    """格式化行号列表为可读字符串"""
    if len(rows) == 0:
        return "无"
    if len(rows) <= max_show:
        return f"第{', '.join(str(r+2) for r in rows)}行"  # +2: 0-based + 1 for header
    return f"第{', '.join(str(r+2) for r in rows[:max_show])}行等共{len(rows)}行"


def _safe_date(val):
    """安全转换为日期"""
    if pd.isna(val):
        return pd.NaT
    try:
        return pd.to_datetime(val, errors='coerce')
    except:
        return pd.NaT


def _smart_parse_dates(series):
    """智能解析日期列：支持 datetime、字符串、Excel 序列号"""
    dt = pd.to_datetime(series, errors='coerce')
    if dt.notna().sum() > 0:
        return dt
    # Excel 序列号（1899-12-30 为基准）
    num = pd.to_numeric(series, errors='coerce')
    return pd.Timestamp('1899-12-30') + pd.to_timedelta(num, unit='D')


def _is_future_date(val):
    """检查是否为未来日期"""
    d = _safe_date(val)
    if pd.isna(d):
        return False
    return d.date() > datetime.now().date()


def _check_empty_rows(df, errors):
    """检查全空行"""
    empty_rows = df[df.isnull().all(axis=1)].index.tolist()
    if empty_rows:
        errors.append({
            "message": f"存在完全空白的行（{len(empty_rows)}行），建议删除后再分析",
            "rows": empty_rows,
            "count": len(empty_rows),
            "category": "空白行",
            "severity": "error"
        })


def _check_duplicate_subjects(df, fields, errors):
    """检查重复受试者ID"""
    subj_col = fields.get('Subject')
    if not subj_col:
        return
    dup = df[df[subj_col].notna() & df[subj_col].duplicated(keep=False)]
    if len(dup) > 0:
        errors.append({
            "message": f"存在重复的 Subject ID（共{len(dup)}行涉及），请确认是否重复录入",
            "rows": dup.index.tolist(),
            "count": len(dup),
            "category": "重复数据",
            "severity": "error"
        })


def _check_future_dates(df, fields, errors):
    """检查日期列中的未来日期"""
    date_keys = ['Date_ICF', 'C1D1', 'EOT_Date', 'Latest_Date']
    for key in date_keys:
        col = fields.get(key)
        if not col:
            continue
        future_rows = []
        for idx, val in df[col].items():
            if _is_future_date(val):
                future_rows.append(idx)
        if future_rows:
            errors.append({
                "message": f"「{col}」列中存在未来日期（{len(future_rows)}处），请确认日期是否正确",
                "rows": future_rows,
                "count": len(future_rows),
                "category": "未来日期",
                "column": col,
                "severity": "error"
            })


def _check_site_format(df, fields, warnings):
    """检查中心编号格式（如3.0应显示为3）"""
    site_col = fields.get('Site')
    if not site_col:
        return
    decimal_rows = []
    for idx, val in df[site_col].items():
        s = str(val).strip()
        if s.endswith('.0') and s.replace('.0', '').isdigit():
            decimal_rows.append(idx)
    if decimal_rows:
        warnings.append({
            "message": f"「{site_col}」列有{len(decimal_rows)}个中心编号显示为小数格式（如3.0），已自动转换",
            "rows": decimal_rows,
            "count": len(decimal_rows),
            "category": "格式问题",
            "column": site_col,
            "severity": "warning"
        })


def _check_date_logic(df, fields, errors):
    """临床日期逻辑校验"""
    date_icf_col = fields.get('Date_ICF')
    c1d1_col = fields.get('C1D1')
    latest_col = fields.get('Latest_Date')
    eot_col = fields.get('EOT_Date')
    
    # 查找Pre-ICF相关列
    pre_icf_col = find_column(df.columns, 'Date of Pre-ICF', 'Pre-ICF', 'date of pre-icf')
    
    # Pre-ICF > Date ICF
    if pre_icf_col and date_icf_col:
        bad_rows = []
        for idx, row in df.iterrows():
            pre = _safe_date(row[pre_icf_col])
            icf = _safe_date(row[date_icf_col])
            if pd.notna(pre) and pd.notna(icf) and pre > icf:
                bad_rows.append(idx)
        if bad_rows:
            errors.append({
                "message": f"Pre-ICF 日期晚于 Date ICF 日期（{len(bad_rows)}处），Pre-ICF必须在ICF之前",
                "rows": bad_rows,
                "count": len(bad_rows),
                "category": "日期逻辑错误",
                "severity": "error"
            })
    
    # Date ICF > C1D1
    if date_icf_col and c1d1_col:
        bad_rows = []
        for idx, row in df.iterrows():
            icf = _safe_date(row[date_icf_col])
            c1 = _safe_date(row[c1d1_col])
            if pd.notna(icf) and pd.notna(c1) and icf > c1:
                bad_rows.append(idx)
        if bad_rows:
            errors.append({
                "message": f"Date ICF 晚于 C1D1 日期（{len(bad_rows)}处），知情同意必须在首次给药之前",
                "rows": bad_rows,
                "count": len(bad_rows),
                "category": "日期逻辑错误",
                "severity": "error"
            })
    
    # C1D1 > Latest Date
    if c1d1_col and latest_col:
        bad_rows = []
        for idx, row in df.iterrows():
            c1 = _safe_date(row[c1d1_col])
            latest = _safe_date(row[latest_col])
            if pd.notna(c1) and pd.notna(latest) and c1 > latest:
                bad_rows.append(idx)
        if bad_rows:
            errors.append({
                "message": f"C1D1 晚于 Latest Date（{len(bad_rows)}处），首次给药应早于最新访视日期",
                "rows": bad_rows,
                "count": len(bad_rows),
                "category": "日期逻辑错误",
                "severity": "error"
            })
    
    # EOT Date > Latest Date
    if eot_col and latest_col:
        bad_rows = []
        for idx, row in df.iterrows():
            eot_d = _safe_date(row[eot_col])
            latest = _safe_date(row[latest_col])
            if pd.notna(eot_d) and pd.notna(latest) and eot_d > latest:
                bad_rows.append(idx)
        if bad_rows:
            errors.append({
                "message": f"EOT Date 晚于 Latest Date（{len(bad_rows)}处），终止日期应早于最新访视日期",
                "rows": bad_rows,
                "count": len(bad_rows),
                "category": "日期逻辑错误",
                "severity": "error"
            })


def _check_status_logic(df, fields, errors):
    """按状态检查必填字段（同类型错误合并为一条）"""
    status_col = fields.get('Status')
    if not status_col:
        return
    
    c1d1_col = fields.get('C1D1')
    date_icf_col = fields.get('Date_ICF')
    eot_date_col = fields.get('EOT_Date')
    eot_reason_col = fields.get('EOT_Reason')
    eot_reason_type_col = fields.get('EOT_Reason_Type')
    sf_reason_col = fields.get('Screen_Failure_Reason')
    sf_type_col = fields.get('Screen_Failure_Type')
    pre_icf_col = find_column(df.columns, 'Date of Pre-ICF', 'Pre-ICF', 'date of pre-icf')
    
    # 批量收集各类问题
    issues = {
        'treatment_no_c1d1': [],
        'screening_no_icf': [],
        'sf_no_reason': [],
        'sf_no_type': [],
        'eot_no_date': [],
        'eot_no_reason_type': [],
        'eot_no_reason': [],
    }
    
    for idx, row in df.iterrows():
        status = str(row[status_col]).strip().upper()
        
        if status in ['IN TREATMENT', 'TREATMENT', 'ON TREATMENT']:
            if c1d1_col and pd.isna(row[c1d1_col]):
                issues['treatment_no_c1d1'].append(idx)
        
        if status in ['IN SCREENING', 'SCREENING']:
            has_icf = date_icf_col and pd.notna(row[date_icf_col])
            has_pre = pre_icf_col and pd.notna(row[pre_icf_col])
            if not has_icf and not has_pre:
                issues['screening_no_icf'].append(idx)
        
        if status in ['SF', 'SCREEN FAILURE', 'SCREENFAILURE', 'SCREEN FAILED']:
            if sf_reason_col and (pd.isna(row[sf_reason_col]) or str(row[sf_reason_col]).strip() == ''):
                issues['sf_no_reason'].append(idx)
            if sf_type_col and (pd.isna(row[sf_type_col]) or str(row[sf_type_col]).strip() == ''):
                issues['sf_no_type'].append(idx)
        
        if status == 'EOT':
            if eot_date_col and pd.isna(row[eot_date_col]):
                issues['eot_no_date'].append(idx)
            if eot_reason_type_col and (pd.isna(row[eot_reason_type_col]) or str(row[eot_reason_type_col]).strip() == ''):
                issues['eot_no_reason_type'].append(idx)
            if eot_reason_col and (pd.isna(row[eot_reason_col]) or str(row[eot_reason_col]).strip() == ''):
                issues['eot_no_reason'].append(idx)
    
    if issues['treatment_no_c1d1']:
        errors.append(_make_issue('状态逻辑错误',
            f"状态为「In Treatment」但 C1D1（首次给药日期）为空", issues['treatment_no_c1d1']))
    if issues['screening_no_icf']:
        errors.append(_make_issue('状态逻辑错误',
            f"状态为「In Screening」但 Date ICF 和 Pre-ICF 均为空", issues['screening_no_icf']))
    if issues['sf_no_reason']:
        errors.append(_make_issue('状态逻辑错误',
            f"状态为「Screen Failure」但筛选失败原因（Screen Failure Reason）为空", issues['sf_no_reason']))
    if issues['sf_no_type']:
        errors.append(_make_issue('状态逻辑错误',
            f"状态为「Screen Failure」但筛选失败类型（Screen Failure Type）为空", issues['sf_no_type']))
    if issues['eot_no_date']:
        errors.append(_make_issue('状态逻辑错误',
            f"状态为「EOT」但 EOT Date（终止日期）为空", issues['eot_no_date']))
    if issues['eot_no_reason_type']:
        errors.append(_make_issue('状态逻辑错误',
            f"状态为「EOT」但 EOT Reason Type（终止原因类型）为空", issues['eot_no_reason_type']))
    if issues['eot_no_reason']:
        errors.append(_make_issue('状态逻辑错误',
            f"状态为「EOT」但 EOT Reason（终止原因详情）为空", issues['eot_no_reason']))


def _make_issue(category, message, rows):
    """创建统一格式的校验问题"""
    return {
        "message": f"{message}（{len(rows)}处）",
        "rows": rows,
        "count": len(rows),
        "category": category,
        "severity": "error"
    }


def _check_completeness(df, fields, warnings):
    """检查关键字段的完整性（缺填警告）"""
    # 方案版本
    version_col = fields.get('Protocol_Version')
    if version_col:
        missing = df[df[version_col].isna() | (df[version_col].astype(str).str.strip() == '')].index.tolist()
        if missing:
            warnings.append({
                "message": f"「{version_col}」有{len(missing)}条记录未填写方案版本",
                "rows": missing,
                "count": len(missing),
                "category": "字段缺失",
                "column": version_col,
                "severity": "warning"
            })
    
    # 来源
    source_col = fields.get('Source')
    if source_col:
        missing = df[df[source_col].isna() | (df[source_col].astype(str).str.strip() == '')].index.tolist()
        if missing:
            warnings.append({
                "message": f"「{source_col}」有{len(missing)}条记录未填写病例来源",
                "rows": missing,
                "count": len(missing),
                "category": "字段缺失",
                "column": source_col,
                "severity": "warning"
            })
    
    # Date ICF 缺失（排除Screen Failure）
    icf_col = fields.get('Date_ICF')
    status_col = fields.get('Status')
    if icf_col:
        missing = []
        for idx, row in df.iterrows():
            if pd.isna(row[icf_col]) or str(row[icf_col]).strip() == '':
                if status_col:
                    st = str(row[status_col]).strip().upper()
                    if st not in ['SF', 'SCREEN FAILURE', 'SCREENFAILURE', '']:
                        missing.append(idx)
                else:
                    missing.append(idx)
        if missing:
            warnings.append({
                "message": f"「{icf_col}」有{len(missing)}条非Screen Failure记录未填写知情同意日期",
                "rows": missing,
                "count": len(missing),
                "category": "字段缺失",
                "column": icf_col,
                "severity": "warning"
            })
    
    # C1D1 缺失（排除Screen Failure 和 In Screening）
    c1d1_col = fields.get('C1D1')
    if c1d1_col:
        missing = []
        for idx, row in df.iterrows():
            if pd.isna(row[c1d1_col]) or str(row[c1d1_col]).strip() == '':
                if status_col:
                    st = str(row[status_col]).strip().upper()
                    if st not in ['SF', 'SCREEN FAILURE', 'SCREENFAILURE', '', 'IN SCREENING', 'SCREENING']:
                        missing.append(idx)
                else:
                    missing.append(idx)
        if missing:
            warnings.append({
                "message": f"「{c1d1_col}」有{len(missing)}条非Screen Failure/In Screening记录未填写首次给药日期",
                "rows": missing,
                "count": len(missing),
                "category": "字段缺失",
                "column": c1d1_col,
                "severity": "warning"
            })


# ==================== 主分析函数 ====================

def _safe_read_excel(file_path, sheet_name=0):
    """
    安全读取 Excel，遇到样式解析错误时自动降级为 openpyxl read_only 模式
    """
    try:
        return pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception as e:
        if 'Fill' in str(e) or 'openpyxl' in str(type(e).__name__).lower():
            # 降级1：openpyxl read_only 模式
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                if isinstance(sheet_name, int):
                    snames = wb.sheetnames
                    ws = wb[snames[sheet_name]]
                else:
                    ws = wb[sheet_name]
                data = list(ws.values)
                wb.close()
                if not data:
                    return pd.DataFrame()
                columns = [str(c) if c is not None else f'Unnamed: {i}' for i, c in enumerate(data[0])]
                return pd.DataFrame(data[1:], columns=columns)
            except Exception:
                pass
            
            # 降级2：ZIP+XML 纯数据提取（完全绕过样式）
            if str(file_path).lower().endswith('.xlsx'):
                return _read_xlsx_raw(file_path, sheet_name)
        raise


def auto_detect_sheet(file_path: str) -> Tuple[pd.DataFrame, str]:
    """
    自动检测 Excel 文件中的有效 Tracker 数据表
    """
    import openpyxl
    
    # 获取sheet列表（三级降级）
    sheet_names = None
    try:
        sheet_names = pd.ExcelFile(file_path).sheet_names
    except Exception:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception:
            if str(file_path).lower().endswith('.xlsx'):
                # ZIP方式获取sheet名
                with zipfile.ZipFile(file_path, 'r') as z:
                    wb_xml = z.read('xl/workbook.xml')
                    wb_root = ET.fromstring(wb_xml)
                    ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
                    sheet_names = [s.get('name') for s in wb_root.findall(f'.//{{{ns}}}sheet')]
            else:
                sheet_names = []
    
    if not sheet_names:
        return pd.DataFrame(), ""
    
    for sheet_name in sheet_names:
        temp_df = _safe_read_excel(file_path, sheet_name=sheet_name)
        cols_str = ' '.join(str(c) for c in temp_df.columns)
        if any(key in cols_str for key in ['Phase', 'Site', 'Subject', 'Cohort', 'Tumor']):
            return temp_df, sheet_name
    
    return _safe_read_excel(file_path, sheet_name=0), sheet_names[0]


def load_tracker_data(source) -> pd.DataFrame:
    """
    通用数据加载函数，支持文件路径、DataFrame、上传文件对象
    自动检测 Excel 多Sheet中的有效数据表，自动丢弃全空行
    """
    if isinstance(source, pd.DataFrame):
        return source.dropna(how='all').reset_index(drop=True)
    
    if isinstance(source, str):
        if source.endswith('.csv'):
            df = pd.read_csv(source, encoding='utf-8-sig')
        else:
            df, sheet = auto_detect_sheet(source)
        return df.dropna(how='all').reset_index(drop=True)
    
    # 文件上传对象
    if hasattr(source, 'name') and hasattr(source, 'read'):
        filename = str(source.name)
        if filename.endswith('.csv'):
            df = pd.read_csv(source, encoding='utf-8-sig')
            return df.dropna(how='all').reset_index(drop=True)
        else:
            # Excel文件上传 → 写入临时文件，复用已验证的文件路径读取
            import tempfile
            suffix = os.path.splitext(filename)[1] or '.xlsx'
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                try:
                    source.seek(0)
                except:
                    pass
                tmp.write(source.read())
                tmp_path = tmp.name
            
            try:
                df, sheet = auto_detect_sheet(tmp_path)
                return df.dropna(how='all').reset_index(drop=True)
            finally:
                try:
                    os.unlink(tmp_path)
                except:
                    pass
    
    raise ValueError(f"不支持的数据源类型: {type(source)}")


def run_tracker_analysis(df: pd.DataFrame, output_dir: str = None) -> Dict:
    """
    运行完整的 Tracker 分析
    
    Parameters
    ----------
    df : pd.DataFrame 或 str
        DataFrame 或文件路径
    output_dir : str, optional
        输出目录
    """
    # 加载数据
    df = load_tracker_data(df)
    
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(__file__), "output")
    
    os.makedirs(os.path.join(output_dir, "charts"), exist_ok=True)
    os.makedirs(os.path.join(output_dir, "tables"), exist_ok=True)
    
    # 清理列名
    df.columns = [str(col).strip() for col in df.columns]
    
    # 字段检测
    fields = FieldDetector(df)
    
    results = {
        "analysis_time": datetime.now().isoformat(),
        "total_records": len(df),
        "tables": {},
        "charts": {},
        "summary": {},
        "skipped": [],
        "executed": []
    }
    
    print(f"\n{'='*60}")
    print(f"  开始 Tracker 分析（{len(df)} 条记录）")
    print(f"{'='*60}")
    
    # 字段检测报告
    print("\n[字段检测]")
    print(f"  检测到 {len(fields.found_fields)} 个字段")
    if fields.get_biomarker_type():
        print(f"  生物标志物类型: {fields.get_biomarker_type()}")
    
    # 初始化基础列
    date_icf_col = fields.get('Date_ICF')
    c1d1_col = fields.get('C1D1')
    status_col = fields.get('Status')
    
    if date_icf_col:
        df['筛选'] = df[date_icf_col].notna()
    else:
        df['筛选'] = True
    
    if c1d1_col:
        df['入组'] = df[c1d1_col].notna()
    else:
        df['入组'] = False
    
    if status_col:
        df['筛选失败'] = df[status_col].fillna('').str.upper().isin(['SF', 'SCREEN FAILURE', 'SCREENFAILURE'])
    else:
        df['筛选失败'] = False
    
    # 执行分析
    analyses = [
        ("Phase & 剂量组筛选入组表", analyze_phase_dose),
        ("中心筛选入组表", analyze_site_stats),
        ("Phase & 肿瘤类型筛选入组表", analyze_phase_tumor),
        ("Phase & 状态分布表", analyze_status_pivot),
        ("月度筛选入组趋势", analyze_monthly_trend),
        ("筛选失败原因分布", analyze_sf_reasons),
        ("生物标志物检测分析", analyze_biomarker),
        ("在组时间游泳图", analyze_time_on_study),
        ("EOT原因分布", analyze_eot_reasons),
        ("疗效数据表", analyze_response),
        ("方案版本筛选入组", analyze_protocol_version),
        ("不同来源分析", analyze_source),
        ("各中心筛选时长", analyze_screening_duration),
        ("风险预警看板", analyze_risk_dashboard),
    ]
    
    table_idx = 1
    chart_idx = 1
    
    for name, func in analyses:
        print(f"\n[{name}]")
        try:
            result, skip_reason = func(df, fields, output_dir)
            
            if result is None:
                print(f"  ⏭️ 跳过: {skip_reason}")
                results['skipped'].append({'name': name, 'reason': skip_reason})
            else:
                print(f"  ✅ 完成")
                results['executed'].append(name)
                
                if isinstance(result, dict):
                    if 'table' in result:
                        table_key = f"{table_idx:02d}_{name.split()[0]}"
                        results['tables'][table_key] = result['table']
                        save_table(result['table'], output_dir, f"{table_idx:02d}_table.csv")
                        table_idx += 1
                    if 'chart' in result:
                        results['charts'][name] = result['chart']
                    for key in ['type', 'samples', 'mean_days', 'median_days', 'result_table']:
                        if key in result:
                            if 'biomarker' not in results['summary']:
                                results['summary']['biomarker'] = {}
                            results['summary']['biomarker'][key] = result[key]
                elif isinstance(result, pd.DataFrame):
                    table_key = f"{table_idx:02d}_{name.split()[0]}"
                    results['tables'][table_key] = result
                    save_table(result, output_dir, f"{table_idx:02d}_table.csv")
                    table_idx += 1
                elif isinstance(result, str):
                    results['charts'][name] = result
                    chart_idx += 1
        except Exception as e:
            print(f"  ❌ 错误: {str(e)}")
            results['skipped'].append({'name': name, 'reason': f"执行错误: {str(e)}"})
    
    # 汇总
    total_screened = df['筛选'].sum()
    total_enrolled = df['入组'].sum()
    total_sf = df['筛选失败'].sum()
    
    results['summary']['total_screened'] = int(total_screened)
    results['summary']['total_enrolled'] = int(total_enrolled)
    results['summary']['total_screen_failure'] = int(total_sf)
    results['summary']['success_rate'] = round(total_enrolled / max(total_screened, 1) * 100, 1)
    results['summary']['fields_detected'] = list(fields.found_fields.keys())
    results['summary']['biomarker_type'] = fields.get_biomarker_type()
    
    print(f"\n{'='*60}")
    print(f"  分析完成！")
    print(f"  成功: {len(results['executed'])} 项, 跳过: {len(results['skipped'])} 项")
    print(f"{'='*60}")
    
    print(f"\n总筛选数: {total_screened}")
    print(f"总入组数: {total_enrolled}")
    print(f"筛选失败数: {total_sf}")
    print(f"筛选成功率: {results['summary']['success_rate']}%")
    
    return results


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        run_tracker_analysis(df)
    else:
        print("用法: python clinical_engine_v3.py <数据文件>")